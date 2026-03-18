"""
Enhanced Sales Accounts Receivable App
Now includes:
- Edit / Delete invoices
- Customer statement report
- Dashboard summary
"""

from datetime import datetime
import os
import pandas as pd
import sys

# ---------- Optional Streamlit Import ----------
STREAMLIT_AVAILABLE = True
try:
    import streamlit as st
except ModuleNotFoundError:
    STREAMLIT_AVAILABLE = False

# ---------- Auto Folder ----------
HOME_DIR = os.path.expanduser("~")
APP_FOLDER = os.path.join(HOME_DIR, "Documents", "AR_App")
os.makedirs(APP_FOLDER, exist_ok=True)

try:
    os.chdir(APP_FOLDER)
except OSError:
    APP_FOLDER = os.getcwd()

DATA_FILE = os.path.join(APP_FOLDER, "data.csv")

COLUMNS = [
"Inv Date","Inv No.","Customer","Shipment","POL","Cont.","FOB SGD","Freight","C&F SGD",
"Pymt rcvd","Balance receivable","Terms (days)","Due Date","Collect Date","Overdue days"
]

# ---------- Data ----------
def load_data():
    try:
        df = pd.read_csv(DATA_FILE)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = None
        df = df.apply(lambda r: calculate_fields(r), axis=1)
        return df
    except:
        return pd.DataFrame(columns=COLUMNS)


def save_data(df):
    df.to_csv(DATA_FILE, index=False)


# ---------- Calculations ----------
def calculate_fields(row):
    try:
        if pd.notnull(row.get("Inv Date")) and pd.notnull(row.get("Terms (days)")):
            inv_date_parsed = pd.to_datetime(row["Inv Date"], format="%d-%b-%y", errors="coerce")
            if pd.notnull(inv_date_parsed):
                row["Due Date"] = (inv_date_parsed + pd.Timedelta(days=int(row["Terms (days)"]))).strftime('%d-%b-%y')

        row["C&F SGD"] = float(row.get("FOB SGD",0)) + float(row.get("Freight",0))

        if pd.notnull(row.get("Pymt rcvd")):
            row["Balance receivable"] = row["C&F SGD"] - row["Pymt rcvd"]

        # ---------- Clear fields if fully paid ----------
        balance = pd.to_numeric(row.get("Balance receivable"), errors="coerce")
        if pd.notnull(balance) and balance == 0:
            row["Due Date"] = ""
            row["Overdue days"] = ""
        else:
            # Calculate overdue only if not fully paid
            if pd.notnull(row.get("Due Date")):
                due_parsed = pd.to_datetime(row["Due Date"], format="%d-%b-%y", errors="coerce")
                if pd.notnull(due_parsed):
                    days = (pd.Timestamp.today().normalize() - due_parsed).days
                    row["Overdue days"] = max(days, 0)
    except:
        pass
    return row


# ---------- Streamlit App ----------
def run_streamlit():

    st.set_page_config(page_title="AR Manager", layout="wide")

    st.title("📊 AR Management System")

    menu = st.sidebar.selectbox("Menu", [
        "Dashboard",
        "Invoice Entry",
        "Edit / Delete",
        "Customer Statement",
        "AR Aging",
        "Closed Invoices",
        "Dashboard",
        "Invoice Entry",
        "Edit / Delete",
        "Customer Statement",
        "AR Aging"
    ])

    df = load_data()

    # Separate Open vs Closed
    df["Balance receivable"] = pd.to_numeric(df["Balance receivable"], errors="coerce")
    open_df = df[df["Balance receivable"].fillna(0) > 0]
    closed_df = df[df["Balance receivable"].fillna(0) == 0]

    # Ensure numeric columns are correct types
    df["Overdue days"] = pd.to_numeric(df.get("Overdue days"), errors="coerce")
    df["Balance receivable"] = pd.to_numeric(df.get("Balance receivable"), errors="coerce")
    df["Pymt rcvd"] = pd.to_numeric(df.get("Pymt rcvd"), errors="coerce")
    df["FOB SGD"] = pd.to_numeric(df.get("FOB SGD"), errors="coerce")
    df["Freight"] = pd.to_numeric(df.get("Freight"), errors="coerce")

    # ---------- Dashboard ----------
    if menu == "Dashboard":
        st.header("Dashboard")

        total = open_df["Balance receivable"].fillna(0).sum()
        overdue = open_df[open_df["Overdue days"].fillna(0) > 0]["Balance receivable"].fillna(0).sum()
        customers = open_df["Customer"].nunique()

        col1, col2, col3 = st.columns(3)
        col1.metric("Total Receivable", f"SGD {total:,.2f}")
        col2.metric("Overdue Amount", f"SGD {overdue:,.2f}")
        col3.metric("Customers", customers)

        st.subheader("Receivable by Customer")
        chart = open_df.groupby("Customer")["Balance receivable"].sum(numeric_only=True)
        st.bar_chart(chart)

        # ---------- Outstanding Invoices Table ----------
        st.subheader("Outstanding Invoices")

        # -------- Filters --------
        col1, col2, col3 = st.columns(3)
        customer_filter = col1.multiselect("Filter by Customer", open_df["Customer"].dropna().unique())
        inv_filter = col2.text_input("Search Invoice No")

        # Date range filter
        open_df["Inv Date Parsed"] = pd.to_datetime(open_df["Inv Date"], format="%d-%b-%y", errors="coerce")
        min_date = open_df["Inv Date Parsed"].min()
        max_date = open_df["Inv Date Parsed"].max()

        date_range = col3.date_input("Invoice Date Range", [min_date, max_date] if pd.notnull(min_date) and pd.notnull(max_date) else [])

        filtered_df = open_df.copy()

        if customer_filter:
            filtered_df = filtered_df[filtered_df["Customer"].isin(customer_filter)]

        if inv_filter:
            filtered_df = filtered_df[filtered_df["Inv No."].astype(str).str.contains(inv_filter, case=False, na=False)]

        if len(date_range) == 2:
            start, end = pd.to_datetime(date_range[0]), pd.to_datetime(date_range[1])
            filtered_df = filtered_df[(filtered_df["Inv Date Parsed"] >= start) & (filtered_df["Inv Date Parsed"] <= end)]

        display_df = filtered_df.copy()

        # Format currency columns
        for col in ["FOB SGD","Freight","C&F SGD","Pymt rcvd","Balance receivable"]:
            if col in display_df.columns:
                display_df[col] = pd.to_numeric(display_df[col], errors="coerce").map(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")

        st.dataframe(display_df)

        # -------- Export --------
        csv = filtered_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download as CSV",
            data=csv,
            file_name="outstanding_invoices.csv",
            mime="text/csv"
        )

                # Excel Export (Formatted)
        import io
        output = io.BytesIO()

        try:
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                filtered_df.to_excel(writer, index=False, sheet_name='Outstanding')
                wb = writer.book
                ws = writer.sheets['Outstanding']

                # Header styling
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

                for col_num, col_name in enumerate(filtered_df.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Auto column width
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Currency formatting
                currency_cols = ["FOB SGD","Freight","C&F SGD","Pymt rcvd","Balance receivable"]
                for col_name in currency_cols:
                    if col_name in filtered_df.columns:
                        col_idx = list(filtered_df.columns).index(col_name) + 1
                        for row in range(2, len(filtered_df) + 2):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.number_format = '#,##0.00'

                # ---------- Totals Row ----------
                total_row = len(filtered_df) + 2
                ws.cell(row=total_row, column=1, value="TOTAL")
                ws.cell(row=total_row, column=1).font = Font(bold=True)

                for col_name in currency_cols:
                    if col_name in filtered_df.columns:
                        col_idx = list(filtered_df.columns).index(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        formula = f"=SUM({col_letter}2:{col_letter}{len(filtered_df)+1})"
                        cell = ws.cell(row=total_row, column=col_idx, value=formula)
                        cell.font = Font(bold=True)
                        cell.number_format = '#,##0.00'

                # ---------- Highlight Overdue Rows ----------
                overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                if "Overdue days" in filtered_df.columns:
                    overdue_idx = list(filtered_df.columns).index("Overdue days") + 1
                    for row in range(2, len(filtered_df) + 2):
                        days_cell = ws.cell(row=row, column=overdue_idx)
                        try:
                            if float(days_cell.value) > 0:
                                for col in range(1, len(filtered_df.columns) + 1):
                                    ws.cell(row=row, column=col).fill = overdue_fill
                        except:
                            pass

            excel_data = output.getvalue()

        except ModuleNotFoundError:
            # Fallback if openpyxl is not installed
            excel_data = filtered_df.to_csv(index=False).encode('utf-8')
            st.warning("openpyxl not installed → exporting as CSV instead of formatted Excel")

        st.download_button(
            label="Download as Excel",
            data=excel_data,
            file_name="outstanding_invoices.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    # ---------- Entry ----------
    elif menu == "Invoice Entry":
        st.header("Invoice Entry")

        with st.form("form"):
            inv_date = st.date_input("Invoice Date")
            inv_no = st.text_input("Invoice No")
            customer = st.text_input("Customer")
            fob = st.number_input("FOB", 0.0)
            freight = st.number_input("Freight", 0.0)
            payment = st.number_input("Payment", 0.0)
            terms = st.number_input("Terms", 0)

            # ---------- Validation ----------
            errors = []
            if not inv_no:
                errors.append("Invoice No is required")
            if not customer:
                errors.append("Customer is required")
            if fob < 0 or freight < 0 or payment < 0:
                errors.append("Amounts cannot be negative")
            if terms < 0:
                errors.append("Terms cannot be negative")

            # ---------- Live Due Date + Format Enforcement ----------
            if inv_date and terms:
                due = pd.to_datetime(inv_date) + pd.Timedelta(days=int(terms))
                st.info(f"Due Date: {due.strftime('%d-%b-%y')}")

            # Enforce display format DD-MMM-YY
            if inv_date:
                formatted_date = pd.to_datetime(inv_date).strftime('%d-%b-%y')
                st.caption(f"Invoice Date format: {formatted_date}")

            if st.form_submit_button("Save"):
                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    row = calculate_fields(pd.Series({
                        "Inv Date": pd.to_datetime(inv_date).strftime('%d-%b-%y'),
                        "Inv No.": inv_no,
                        "Customer": customer,
                        "FOB SGD": fob,
                        "Freight": freight,
                        "Pymt rcvd": payment,
                        "Terms (days)": terms
                    }))

                    df = pd.concat([df, pd.DataFrame([row])])
                    save_data(df)
                    st.success("Saved")


    # ---------- Edit/Delete ----------
    elif menu == "Edit / Delete":
        st.header("Edit / Delete Invoices")

        if df.empty:
            st.warning("No data")
            return

        # Build detailed invoice labels
        display_df = df.copy()
        # Robust date parsing (avoids warning for mixed formats)
        display_df["Inv Date"] = pd.to_datetime(display_df["Inv Date"], format="%d-%b-%y", errors="coerce").dt.strftime('%d-%b-%y')
        display_df["Label"] = (
            display_df["Inv No."].fillna("(No Inv No)").astype(str)
            + " | "
            + display_df["Customer"].fillna("(No Customer)").astype(str)
            + " | "
            + display_df["Inv Date"].astype(str)
        )

        selected_label = st.selectbox("Select Invoice", display_df["Label"])

        # Get index from selected label
        idx = display_df[display_df["Label"] == selected_label].index[0]
        row = df.loc[idx]

        new_customer = st.text_input("Customer", row["Customer"])
        new_payment = st.number_input("Payment", value=float(row.get("Pymt rcvd",0)))

        if st.button("Update"):
            df.at[idx, "Customer"] = new_customer
            df.at[idx, "Pymt rcvd"] = new_payment
            df = df.apply(lambda r: calculate_fields(r), axis=1)
            save_data(df)
            st.success("Updated")

        if st.button("Delete"):
            df = df.drop(idx)
            save_data(df)
            st.warning("Deleted")


    # ---------- Customer Statement ----------
    elif menu == "Customer Statement":
        st.header("Customer Statement")

        customer = st.selectbox("Select Customer", open_df["Customer"].dropna().unique())

        cust_df = open_df[open_df["Customer"] == customer].copy()
        cust_df["Balance receivable"] = pd.to_numeric(cust_df["Balance receivable"], errors="coerce")

        # Format currency columns for display
        display_df = cust_df.copy()
        for col in ["FOB SGD","Freight","C&F SGD","Pymt rcvd","Balance receivable"]:
            if col in display_df.columns:
                display_df[col] = pd.to_numeric(display_df[col], errors="coerce").map(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")

        st.dataframe(display_df)

        total = cust_df["Balance receivable"].fillna(0).sum()
        st.metric("Total Outstanding", f"SGD {total:,.2f}")


    # ---------- Aging ----------
    elif menu == "AR Aging":
        st.header("AR Aging")

        df["Balance receivable"] = pd.to_numeric(df["Balance receivable"], errors="coerce").fillna(0)

        # Calculate overdue days fresh from Due Date
        df["Due Date Parsed"] = pd.to_datetime(df["Due Date"], format="%d-%b-%y", errors="coerce")
        today = pd.Timestamp.today().normalize()
        df["Overdue days"] = (today - df["Due Date Parsed"]).dt.days
        df["Overdue days"] = df["Overdue days"].apply(lambda x: x if x > 0 else 0)

        # ---------- Standard Aging Buckets ----------
        aging = {
            "Current": df[df["Overdue days"] == 0]["Balance receivable"].sum(),
            "1-30": df[(df["Overdue days"] > 0) & (df["Overdue days"] <= 30)]["Balance receivable"].sum(),
            "31-60": df[(df["Overdue days"] > 30) & (df["Overdue days"] <= 60)]["Balance receivable"].sum(),
            "61-90": df[(df["Overdue days"] > 60) & (df["Overdue days"] <= 90)]["Balance receivable"].sum(),
            "90+": df[df["Overdue days"] > 90]["Balance receivable"].sum()
        }

        aging_df = pd.DataFrame(aging.items(), columns=["Bucket","Amount"])
        aging_df["Amount"] = aging_df["Amount"].fillna(0).map(lambda x: f"{x:,.2f}")

        st.subheader("Overall Aging")
        st.table(aging_df)
        st.bar_chart(pd.DataFrame(list(aging.values()), index=list(aging.keys()), columns=["Amount"]))

        # ---------- Aging by Customer ----------
        st.subheader("Aging by Customer")

        def bucketize(days):
            if days == 0:
                return "Current"
            elif days <= 30:
                return "1-30"
            elif days <= 60:
                return "31-60"
            elif days <= 90:
                return "61-90"
            else:
                return "90+"

        df["Bucket"] = df["Overdue days"].apply(bucketize)

        cust_aging = df.pivot_table(
            index="Customer",
            columns="Bucket",
            values="Balance receivable",
            aggfunc="sum",
            fill_value=0
        )

        cust_aging = cust_aging.reset_index()

        # Format only numeric columns to avoid errors
        numeric_cols = cust_aging.select_dtypes(include="number").columns
        st.dataframe(cust_aging.style.format({col: "{:,.2f}" for col in numeric_cols}))

        # ---------- Alerts ----------
        st.subheader("Alerts (Over 90 Days)")
        alerts_df = open_df[open_df["Overdue days"] > 90].copy()

        if alerts_df.empty:
            st.success("No critical overdue invoices 🎉")
        else:
            alerts_df = alerts_df.sort_values(by="Overdue days", ascending=False)
            st.error(f"{len(alerts_df)} invoices over 90 days!")
            st.dataframe(alerts_df[["Inv No.","Customer","Due Date","Overdue days","Balance receivable"]])

        # ---------- Aging Trend ----------
        st.subheader("Aging Trend (Monthly)")

        df["Inv Date Parsed"] = pd.to_datetime(df["Inv Date"], format="%d-%b-%y", errors="coerce")
        df["Month"] = df["Inv Date Parsed"].dt.to_period("M").astype(str)

        trend = open_df.groupby("Month")["Balance receivable"].sum().sort_index()

        st.line_chart(trend)


# ---------- Closed Invoices ----------
    elif menu == "Closed Invoices":
        st.header("Closed (Paid) Invoices")

        if closed_df.empty:
            st.info("No closed invoices")
        else:
            display_df = closed_df.copy()

            for col in ["FOB SGD","Freight","C&F SGD","Pymt rcvd","Balance receivable"]:
                if col in display_df.columns:
                    display_df[col] = pd.to_numeric(display_df[col], errors="coerce").map(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")

            st.dataframe(display_df)

            total_closed = pd.to_numeric(closed_df["C&F SGD"], errors="coerce").fillna(0).sum()
            st.metric("Total Closed Value", f"SGD {total_closed:,.2f}")


# ---------- Run ----------
if __name__ == "__main__":
    if STREAMLIT_AVAILABLE:
        run_streamlit()
    else:
        print("Please install streamlit to run the app UI")
